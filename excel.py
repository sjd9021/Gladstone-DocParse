import openpyxl
from tess import our_ref, branch, insurers, date, total_amount, policy
import pandas as pd
from openpyxl.styles import Font, Alignment
c = 0
df = pd.read_excel('gal.xlsx')
a = df.iloc[:, 1]
serial = df['SL NO'].iloc[-1]
serial = serial + 1
for i in a:
    if our_ref == i:
        c = 1

if c == 0:
    wb = openpyxl.load_workbook('gal.xlsx')
    ws = wb.active
    newRowLocation = ws.max_row + 1
    ws.cell(column=1, row=newRowLocation, value=serial)
    ws.cell(column=2, row=newRowLocation, value=our_ref)
    ws.cell(column=4, row=newRowLocation, value=insurers)
    ws.cell(column=5, row=newRowLocation, value=branch)
    ws.cell(column=6, row=newRowLocation, value=policy)
    ws.cell(column=7, row=newRowLocation, value=date)
    ws.cell(column=9, row=newRowLocation, value=total_amount)
    ws.cell(column=10, row=newRowLocation, value=total_amount)
    ws.cell(column=13, row=newRowLocation, value='OPEN')
    ft1 = Font(name='Arial', size=12)
    for colNum in range(1, ws.max_column + 1):
        ws.cell(row=newRowLocation, column=colNum).font = ft1
        ws.cell(row=newRowLocation, column=colNum).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    wb.save('gal.xlsx')
    wb.close()
